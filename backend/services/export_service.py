"""
Export service — generates CSV and XLSX from analysis results.
Uses only the stdlib csv module + openpyxl (no pandas dependency).
"""
from __future__ import annotations

import csv
import io
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from backend.models.schemas import (
    AnalysisResult,
    FamilyEntry,
    LemmaEntry,
    VocabEntry,
    WordEntry,
)


# ── Helpers ───────────────────────────────────────────────────────────────────

def _header_style(cell, color: str = "4472C4"):
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=color)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _auto_width(ws, min_w: int = 8, max_w: int = 40):
    for col in ws.columns:
        length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(
            max_w, max(min_w, length + 2)
        )


# ── CSV export ────────────────────────────────────────────────────────────────

def _vocab_rows(result: AnalysisResult, selected_only: bool = False) -> list:
    rows = result.vocab_table or []
    if selected_only:
        rows = [v for v in rows if v.selected is not False]
    return rows


def to_csv(result: AnalysisResult, selected_only: bool = False) -> bytes:
    buf = io.StringIO()
    w = csv.writer(buf)

    # Sheet: word_table
    w.writerow(["=== RAW WORD FREQUENCY ==="])
    w.writerow(["surface", "lemma", "pos", "family_id",
                "body_count", "stem_count", "option_count", "total_count", "score"])
    for e in result.word_table:
        w.writerow([e.surface, e.lemma, e.pos, e.family_id or "",
                    e.body_count, e.stem_count, e.option_count, e.total_count, e.score])

    w.writerow([])
    # Sheet: lemma_table
    w.writerow(["=== LEMMA GROUPS ==="])
    w.writerow(["lemma", "pos", "family_id", "surface_forms",
                "body_count", "stem_count", "option_count", "total_count", "score"])
    for e in result.lemma_table:
        w.writerow([e.lemma, e.pos, e.family_id or "", " | ".join(e.surface_forms),
                    e.body_count, e.stem_count, e.option_count, e.total_count, e.score])

    w.writerow([])
    # Sheet: family_table
    w.writerow(["=== WORD FAMILIES ==="])
    w.writerow(["family_id", "members",
                "body_count", "stem_count", "option_count", "total_count", "score"])
    for e in result.family_table:
        w.writerow([e.family_id, " | ".join(e.members),
                    e.body_count, e.stem_count, e.option_count, e.total_count, e.score])

    w.writerow([])
    # Sheet: vocab_table
    vocab = _vocab_rows(result, selected_only)
    if vocab:
        w.writerow(["=== VOCABULARY LIST ==="])
        w.writerow(["headword", "lemma", "family", "pos",
                    "chinese_meaning", "english_definition", "example_sentence",
                    "notes", "word_level", "selected",
                    "body_count", "stem_count", "option_count",
                    "total_count", "score", "source"])
        for e in vocab:
            w.writerow([
                e.headword, e.lemma, e.family or "", e.pos or "",
                e.chinese_meaning or "", e.english_definition or "",
                e.example_sentence or "", e.notes or "",
                e.word_level or "", "Y" if e.selected is not False else "N",
                e.body_count, e.stem_count, e.option_count, e.total_count,
                e.score, e.source,
            ])

    return buf.getvalue().encode("utf-8-sig")   # BOM for Excel compatibility


# ── XLSX export ───────────────────────────────────────────────────────────────

_SHEET_HEADERS = {
    "Raw Freq": ["surface", "lemma", "pos", "family_id",
                 "body_count", "stem_count", "option_count", "total_count", "score"],
    "Lemma Groups": ["lemma", "pos", "family_id", "surface_forms",
                     "body_count", "stem_count", "option_count", "total_count", "score"],
    "Word Families": ["family_id", "members",
                      "body_count", "stem_count", "option_count", "total_count", "score"],
    "Vocabulary": ["headword", "lemma", "family", "pos",
                   "chinese_meaning", "english_definition", "example_sentence",
                   "notes", "body_count", "stem_count", "option_count",
                   "total_count", "score", "source"],
}

_TAB_COLORS = {
    "Raw Freq": "4472C4",
    "Lemma Groups": "70AD47",
    "Word Families": "ED7D31",
    "Vocabulary": "9B59B6",
}


def to_xlsx(result: AnalysisResult, selected_only: bool = False) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)   # remove default sheet

    # ── Raw Freq ──
    ws = wb.create_sheet("Raw Freq")
    ws.sheet_properties.tabColor = _TAB_COLORS["Raw Freq"]
    ws.freeze_panes = "A2"
    headers = _SHEET_HEADERS["Raw Freq"]
    ws.append(headers)
    for cell in ws[1]:
        _header_style(cell, _TAB_COLORS["Raw Freq"])
    for e in result.word_table:
        ws.append([e.surface, e.lemma, e.pos, e.family_id or "",
                   e.body_count, e.stem_count, e.option_count, e.total_count, e.score])
    _auto_width(ws)

    # ── Lemma Groups ──
    ws = wb.create_sheet("Lemma Groups")
    ws.sheet_properties.tabColor = _TAB_COLORS["Lemma Groups"]
    ws.freeze_panes = "A2"
    ws.append(_SHEET_HEADERS["Lemma Groups"])
    for cell in ws[1]:
        _header_style(cell, _TAB_COLORS["Lemma Groups"])
    for e in result.lemma_table:
        ws.append([e.lemma, e.pos, e.family_id or "", " | ".join(e.surface_forms),
                   e.body_count, e.stem_count, e.option_count, e.total_count, e.score])
    _auto_width(ws)

    # ── Word Families ──
    ws = wb.create_sheet("Word Families")
    ws.sheet_properties.tabColor = _TAB_COLORS["Word Families"]
    ws.freeze_panes = "A2"
    ws.append(_SHEET_HEADERS["Word Families"])
    for cell in ws[1]:
        _header_style(cell, _TAB_COLORS["Word Families"])
    for e in result.family_table:
        ws.append([e.family_id, " | ".join(e.members),
                   e.body_count, e.stem_count, e.option_count, e.total_count, e.score])
    _auto_width(ws)

    # ── Vocabulary ──
    vocab = _vocab_rows(result, selected_only)
    if vocab:
        ws = wb.create_sheet("Vocabulary")
        ws.sheet_properties.tabColor = _TAB_COLORS["Vocabulary"]
        ws.freeze_panes = "A2"
        headers = _SHEET_HEADERS["Vocabulary"] + ["word_level", "selected"]
        ws.append(headers)
        for cell in ws[1]:
            _header_style(cell, _TAB_COLORS["Vocabulary"])
        for e in vocab:
            ws.append([
                e.headword, e.lemma, e.family or "", e.pos or "",
                e.chinese_meaning or "", e.english_definition or "",
                e.example_sentence or "", e.notes or "",
                e.body_count, e.stem_count, e.option_count, e.total_count,
                e.score, e.source,
                e.word_level or "", "Y" if e.selected is not False else "N",
            ])
        _auto_width(ws)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Anki deck export ──────────────────────────────────────────────────────────

def _anki_field(text: object) -> str:
    """Sanitise a value for a single Anki TSV field.

    Tab and carriage-return are the column/row separators in a TSV import, so
    they must never appear inside a field; newlines become ``<br>`` (the import
    runs with ``#html:true``). Returns "" for None/empty.
    """
    if not text:
        return ""
    s = str(text)
    return s.replace("\t", " ").replace("\r", " ").replace("\n", "<br>").strip()


def _anki_tags(item: dict) -> str:
    """Build a space-separated Anki tag string from a library row.

    Anki tags are space-delimited, so any space inside a tag is collapsed to an
    underscore. The word's CEFR/level is added as an extra tag for easy
    filtering inside Anki.
    """
    tags: list[str] = []
    for raw in (item.get("tags") or "").split(","):
        t = raw.strip().replace(" ", "_")
        if t:
            tags.append(t)
    level = item.get("cefr_level") or item.get("word_level")
    if level:
        tags.append(str(level).strip().replace(" ", "_"))
    # Dedupe while preserving order.
    seen: set[str] = set()
    ordered = [t for t in tags if not (t in seen or seen.add(t))]
    return " ".join(ordered)


def to_anki_tsv(items: list[dict]) -> bytes:
    """Render the user's library as an Anki-importable tab-separated deck.

    Layout (one note per saved word):
      * **Front** — the headword (plus part-of-speech, if known)
      * **Back**  — Chinese meaning, English definition and an example sentence,
                    laid out with simple HTML so the card is readable
      * **Tags**  — the word's library tags + its CEFR/level

    The leading ``#`` directives are understood by Anki 2.1+ so the columns and
    tags map automatically on import — no manual field assignment needed.
    """
    out = io.StringIO()
    out.write("#separator:tab\n")
    out.write("#html:true\n")
    out.write("#columns:Front\tBack\tTags\n")
    out.write("#tags column:3\n")

    for item in items:
        headword = _anki_field(item.get("headword") or item.get("lemma"))
        if not headword:
            continue
        pos = _anki_field(item.get("pos"))
        front = f"{headword} <small>({pos})</small>" if pos else headword

        back_parts: list[str] = []
        meaning = _anki_field(item.get("chinese_meaning"))
        if meaning:
            back_parts.append(meaning)
        definition = _anki_field(item.get("english_definition"))
        if definition:
            back_parts.append(f"<i>{definition}</i>")
        example = _anki_field(item.get("example_sentence"))
        if example:
            back_parts.append(f"<br><span style='color:#666'>{example}</span>")
        notes = _anki_field(item.get("notes"))
        if notes:
            back_parts.append(f"<br>{notes}")
        back = "<br>".join(back_parts)

        out.write(f"{front}\t{back}\t{_anki_tags(item)}\n")

    return out.getvalue().encode("utf-8")
